# logger_bot.py
import os
import sys
import json
import threading
import concurrent.futures
import asyncio
from datetime import timezone, datetime
from typing import Callable, Optional

import discord
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------
# Paths
# ---------------------------

def get_app_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

APP_DIR = get_app_dir()
CONFIG_PATH = os.path.join(APP_DIR, "config.json")
OUT_DIR = os.path.join(APP_DIR, "logs")
XLSX_PATH = os.path.join(OUT_DIR, "channel-log.xlsx")
SHEET_NAME = "Messages"

HEADERS = ["Time (Local)", "Author", "Message", "Attachments", "MessageID"]
COLUMN_WIDTHS = [22, 24, 80, 45, 18]  # MessageID hidden
BACKFILL_SAVE_EVERY = 50

# ---------------------------
# Runtime state
# ---------------------------

_client: Optional[discord.Client] = None
_bot_thread: Optional[threading.Thread] = None

TOKEN: Optional[str] = None
LOG_CHANNEL_ID: Optional[int] = None
LOGGING_ENABLED: bool = False

# Status for GUI
_STATUS_LOCK = threading.Lock()
_STATUS = {
    "bot_state": "OFFLINE",        # OFFLINE | CONNECTING | ONLINE | ERROR
    "logging_state": "OFF",        # OFF | ON
    "last_message_time": None,     # ISO string
    "messages_logged": 0,
    "last_error": None,
    "config_autostart": False,
}

_status_listeners: list[Callable[[dict], None]] = []

def _emit_status(note: str | None = None):
    """Notify listeners that status changed."""
    snapshot = get_status()
    if note:
        snapshot["note"] = note
    for cb in list(_status_listeners):
        try:
            cb(snapshot)
        except Exception:
            pass

def subscribe_status(callback: Callable[[dict], None]):
    _status_listeners.append(callback)

def get_status() -> dict:
    with _STATUS_LOCK:
        return dict(_STATUS)

def _set_status(**kwargs):
    with _STATUS_LOCK:
        _STATUS.update(kwargs)

# ---------------------------
# Config (friendly errors)
# ---------------------------

def ensure_config_exists() -> bool:
    if os.path.exists(CONFIG_PATH):
        return True
    os.makedirs(APP_DIR, exist_ok=True)
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(
            {
                "token": "PUT_BOT_TOKEN_HERE",
                "log_channel_id": 123456789012345678,
                "autostart": False
            },
            f,
            indent=2,
        )
    return False

def load_raw_config() -> dict:
    existed = ensure_config_exists()
    if not existed:
        raise ValueError(
            "SETUP: Created config.json next to the app.\n"
            "Edit it and replace token + log_channel_id, then click Start again."
        )
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        raise ValueError("config.json is invalid JSON. Delete it and rerun to regenerate.")

def save_config(update: dict):
    cfg = load_raw_config()
    cfg.update(update)
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2)

def load_config() -> tuple[str, int, bool]:
    cfg = load_raw_config()

    token = str(cfg.get("token", "")).strip()
    ch_raw = cfg.get("log_channel_id", 0)
    autostart = bool(cfg.get("autostart", False))

    _set_status(config_autostart=autostart)

    if not token or token == "PUT_BOT_TOKEN_HERE":
        raise ValueError("SETUP: config.json still has the placeholder token. Paste the real bot token.")

    try:
        channel_id = int(ch_raw)
    except Exception:
        raise ValueError("log_channel_id must be a number (channel ID).")

    if channel_id <= 0:
        raise ValueError("log_channel_id looks invalid.")

    return token, channel_id, autostart

def set_autostart(value: bool):
    save_config({"autostart": bool(value)})
    _set_status(config_autostart=bool(value))
    _emit_status("Auto-start setting updated.")

# ---------------------------
# Excel helpers
# ---------------------------

def _apply_sheet_formatting(ws):
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"

    for i, w in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.column_dimensions["E"].hidden = True

def ensure_workbook():
    os.makedirs(OUT_DIR, exist_ok=True)

    if os.path.exists(XLSX_PATH):
        wb = load_workbook(XLSX_PATH)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.create_sheet(SHEET_NAME)

        if ws.max_row == 1 and ws["A1"].value is None:
            ws.append(HEADERS)
            _apply_sheet_formatting(ws)
            wb.save(XLSX_PATH)
            return wb, ws

        existing_headers = [ws.cell(row=1, column=c).value for c in range(1, 6)]
        if existing_headers[:4] == HEADERS[:4] and existing_headers[4] != "MessageID":
            ws.cell(row=1, column=5).value = "MessageID"

        _apply_sheet_formatting(ws)
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    _apply_sheet_formatting(ws)
    wb.save(XLSX_PATH)
    return wb, ws

def get_last_logged_message_id(ws):
    if ws.max_row < 2:
        return None
    for r in range(ws.max_row, 1, -1):
        val = ws.cell(row=r, column=5).value
        if val:
            try:
                return int(str(val).strip())
            except ValueError:
                continue
    return None

def format_local_time_from_message(message: discord.Message) -> str:
    created = message.created_at
    if created.tzinfo is None:
        created = created.replace(tzinfo=timezone.utc)
    local_dt = created.astimezone()
    return local_dt.strftime("%Y-%m-%d %I:%M:%S %p")

def append_message_row(ws, message: discord.Message):
    local_time = format_local_time_from_message(message)
    author = getattr(message.author, "display_name", None) or message.author.name
    content = (message.content or "").strip()
    attachments = " ".join(a.url for a in message.attachments)
    msg_id = str(message.id)

    ws.append([local_time, author, content, attachments, msg_id])

    row_idx = ws.max_row
    ws.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=row_idx, column=4).alignment = Alignment(wrap_text=True, vertical="top")

async def backfill_history(channel, wb, ws):
    last_id = get_last_logged_message_id(ws)
    after_obj = discord.Object(id=last_id) if last_id else None

    count = 0
    async for msg in channel.history(limit=None, oldest_first=True, after=after_obj):
        if msg.author.bot:
            continue
        append_message_row(ws, msg)
        count += 1
        if count % BACKFILL_SAVE_EVERY == 0:
            wb.save(XLSX_PATH)
            _emit_status(f"Backfill progress: {count} messages...")

    if count:
        wb.save(XLSX_PATH)
    return count

# ---------------------------
# Bot wiring
# ---------------------------

def _make_client() -> discord.Client:
    intents = discord.Intents.default()
    intents.guilds = True
    intents.messages = True
    intents.message_content = True

    client = discord.Client(intents=intents)

    @client.event
    async def on_ready():
        _set_status(bot_state="ONLINE", last_error=None)
        _emit_status("Bot connected.")

    @client.event
    async def on_disconnect():
        _set_status(bot_state="OFFLINE")
        _emit_status("Bot disconnected.")

    @client.event
    async def on_message(message: discord.Message):
        global LOGGING_ENABLED, LOG_CHANNEL_ID

        if not LOGGING_ENABLED:
            return
        if message.author.bot:
            return
        if LOG_CHANNEL_ID is None or message.channel.id != LOG_CHANNEL_ID:
            return

        wb, ws = ensure_workbook()
        last_id = get_last_logged_message_id(ws)
        if last_id == message.id:
            return

        append_message_row(ws, message)
        wb.save(XLSX_PATH)

        # Update status counters
        with _STATUS_LOCK:
            _STATUS["messages_logged"] += 1
            _STATUS["last_message_time"] = datetime.now().astimezone().isoformat()
        _emit_status()

    return client

def is_bot_running() -> bool:
    return _client is not None and _client.is_ready()

def get_logs_dir() -> str:
    os.makedirs(OUT_DIR, exist_ok=True)
    return OUT_DIR

def start_bot_background():
    """
    Start Discord client in a background thread (only once).
    Raises ValueError for config issues.
    """
    global _client, _bot_thread, TOKEN, LOG_CHANNEL_ID

    if _bot_thread and _bot_thread.is_alive():
        return

    token, channel_id, _autostart = load_config()
    TOKEN, LOG_CHANNEL_ID = token, channel_id

    _set_status(bot_state="CONNECTING", last_error=None)
    _emit_status("Starting bot...")

    def runner():
        global _client
        _client = _make_client()
        try:
            _client.run(TOKEN)
        except discord.LoginFailure:
            _set_status(bot_state="ERROR", last_error="Invalid Discord token.")
            _emit_status("Invalid Discord token (LoginFailure).")
        except Exception as e:
            _set_status(bot_state="ERROR", last_error=str(e))
            _emit_status(f"Bot crashed: {e}")

    _bot_thread = threading.Thread(target=runner, daemon=True)
    _bot_thread.start()

async def _start_logging_async() -> str:
    global LOGGING_ENABLED, LOG_CHANNEL_ID, _client

    if _client is None or not _client.is_ready():
        return "Bot is not connected yet. Wait a moment and try again."

    if LOGGING_ENABLED:
        return "Logging is already ON."

    channel = _client.get_channel(LOG_CHANNEL_ID)
    if channel is None:
        return "Bot can't see the configured channel. Check permissions + channel ID."

    wb, ws = ensure_workbook()
    try:
        _emit_status("Syncing history (backfill)...")
        appended = await backfill_history(channel, wb, ws)
    except discord.Forbidden:
        return "Missing permission: View Channel + Read Message History."
    except discord.HTTPException as e:
        return f"Discord error during backfill: {e}"

    LOGGING_ENABLED = True
    _set_status(logging_state="ON")
    _emit_status(f"Logging started. Backfilled {appended} messages.")
    return f"Logging started. Backfilled {appended} messages.\nLog: {XLSX_PATH}"

def start_logging(timeout_seconds: float = 120.0) -> str:
    """
    Start logging by scheduling the coroutine on the Discord client's event loop.
    This avoids cross-event-loop discord.py errors when called from the GUI thread.
    """
    global _client

    if _client is None:
        return "SETUP: Bot is not running yet. Click Start again."

    loop = getattr(_client, "loop", None)
    if loop is None or not loop.is_running():
        return "SETUP: Bot is not connected yet. Wait a moment and try again."

    future = asyncio.run_coroutine_threadsafe(_start_logging_async(), loop)
    try:
        return future.result(timeout=timeout_seconds)
    except concurrent.futures.TimeoutError:
        return "ERROR: Start/backfill timed out. Try again (or reduce history)."
    except Exception as e:
        return f"ERROR: {e}"

def stop_logging() -> str:
    global LOGGING_ENABLED
    if not LOGGING_ENABLED:
        _set_status(logging_state="OFF")
        _emit_status("Logging already OFF.")
        return "Logging is already OFF."
    LOGGING_ENABLED = False
    _set_status(logging_state="OFF")
    _emit_status("Logging stopped.")
    return "Logging stopped."